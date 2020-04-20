import openpyxl as pyxl
import pandas as pd

#엑셀 파일(WorkBook) 읽기

time = list()
pol_data = list()
#name = list()
date = list()
tot_date = list()
O3 = list()
NO = list()
CO = list()
SO = list()
PM10 = list()
PM25 = list()

#관측 지점 목록
obs_site = ['광복동', '장림동', '학장동', '덕천동', '연산동', '대연동', '청룡동', 
    '전포동', '태종대','기장읍','대저동','부곡동','광안동','명장동','녹산동','용수리',
        '좌동','수정동','대신동','온천동','초량동']

#일(day) 수 초기화
num_day = 0

#24시간 list를 미리 만들기
#10 이하 시간에는 숫자'0'추가, 10 이상에는 그냥 time 리스트에 저장
for t in range(1,25):
    if t < 10 :
        time.append('0'+ str(t))
    else:
        time.append(t)
# time.append((r[t].value)[:-1])

for yr in range(2010, 2012):# 2017):
    for mn in range(1, 13): #  13):

        file_name = "data/측정기록지("+str(yr)+"년"+str(mn)+"월).xlsx"

        print(yr, mn, file_name)

        wb = pyxl.load_workbook(file_name)

        test = wb.worksheets[0] #시트 순서별로 불러오기

        pol_data = list()
        # name = list()
        # time = list()
        date = list()

        #open한 엑셀 파일의 line 수 count
        check_line = 0

        for r in test.rows:
            check_line = check_line + 1

            if r[0].value == '측 정 소: ' :
                continue
            elif r[0].value == '측정항목: ' :
                name = r[2].value
                print(name)
            elif r[0].value == '측정년월: ' :
                year = (r[1].value)[0:5]
                #month = int((r[2].value)[0:2])
                month = (r[2].value)[0:2]
                print(year, month)
            elif r[0].value == '구 분' :
                continue
            elif r[0].value == '최소' :
                continue                
            elif r[0].value == '최대' :
                if name == 'O₃':
                    O3 = O3 + pol_data
                    pol_data = list()
                    date = list()
                elif name == 'SO₂':
                    SO = SO + pol_data
                    pol_data = list()
                    date = list()
                elif name == 'CO':
                    CO = CO + pol_data
                    pol_data = list()
                    date = list()
                elif name == 'NO₂':
                    NO = NO + pol_data
                    pol_data = list()
                    date = list()
                elif name == 'PM-10':
                    if test.max_row == check_line + 1:
                        print(test.max_row, check_line)
                        PM10 = PM10 + pol_data
                        tot_date = tot_date + date
                        #PM25 리스트에 일수만큼의 null값(-999)을 넣어라 (일수*24시간)
                        for x in range(num_day*24):
                            PM25.append(-999)
                    else:
                        PM10 = PM10 + pol_data
                        print(test.max_row, check_line)
                        pol_data = list()
                        date = list()
                elif name == 'PM-2.5':
                    PM25 = PM25 + pol_data
                    tot_date = tot_date + date
                    break
                else:
                    pol_data = list()
                    date = list()
            elif r[0].value == '평균' :
                num_day = 0
                continue
            elif r[0].value == None :
                continue
            else : #오염 자료가 있는 1일부터 말일까지 읽기
                #날짜 중 '일'에 숫자 0 붙히기
                num_day = num_day + 1

                if int((r[0].value)[:-1]) < 10 :
                    day = '0'+ (r[0].value)[:-1]
                else :
                    day = (r[0].value)[:-1]

                #각 오염 농도값을 pol_data 리스트에 추가하기
                for i in range(1,25):
                    obj = r[i].value
                    pol_data.append(obj)
                    #pd.Series(pol_data).fillna('-999')
                    
                    # if (r[i].value).isnull() == None:
                    #     r[i] = 'Null'
                    #     pol_data.append(r[i].value)
                    # else:
                    #     pol_data.append(r[i].value)
                                        
                    date.append(str(year)+str(month)+str(day)+str(time[i-1]))
                    # print(year, month, day, time[i-1], pol_data[i-1])


df = pd.DataFrame({'Date':pd.Series(tot_date),'O₃':pd.Series(O3), 'SO₂':pd.Series(SO), 'CO':pd.Series(CO), 'NO₂':pd.Series(NO),'PM-10':pd.Series(PM10), 'PM-2.5':pd.Series(PM25)})

#d = {'O₃':O3, 'SO₂':SO, 'CO':CO, 'NO₂':NO,'PM-10':PM10, 'PM-2.5':PM2.5}
#df = pd.DataFrame(d)
#df.insert(0,'date', date)

print(df)

# index를 date로 바꾸려면 아래 문장 사용해야함.
#df.set_index('date', inplace=True)
#print(df)

df.to_excel('result_test.xlsx',sheet_name='Sheet1')

