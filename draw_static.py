import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl as pyxl
from matplotlib import font_manager, rc

import datetime
from dateutil.parser import parse

#그림에서 한글 깨짐 해결 (컴퓨터에 저장되어있는 한글 font 불러와서 인식시킴)
plt.rcParams["axes.unicode_minus"] = False
path = r'C:\Windows\Fonts\NanumGothic-Regular.ttf'
font_name = font_manager.FontProperties(fname=path).get_name()
rc("font", family = font_name)

################

df = pd.read_excel("test_boxplot.xlsx")

fig, ax = plt.subplots(figsize=(8,3))

ax.set_xscale('log')

sns.boxenplot(x='concentration', y='species', data = df, palette='vlag')

sns.swarmplot(x='concentration', y='species', data = df, size=2, color='.3', linewidth=0)

ax.xaxis.grid(True)
ax.set(ylabel="")
sns.despine(trim=True, left=True)

plt.show()

quit()

df = pd.read_excel("result_test.xlsx", "광복동")

cols = df.columns[[6,7]]


#line
#sns.lineplot(x = df['Date'], y=df['PM-10'])

#heatmap
# corrdata = df[cols].corr(method='pearson').round(2)

# sns.heatmap(data = corrdata, annot=True, annot_kws={"size":9})

#pairplot
# g = sns.pairplot(df[cols])
# g.map_lower(sns.kdeplot)
# g.map_upper(sns.scatterplot)
# g.map_diag(sns.kdeplot, lw=3)

#sns.relplot(x = "Date", y = "PM-10", data = df)

#scatter plot
#sns.regplot(x=df['PM-10'], y=df['PM-2.5'])#, dropna = False)

#KDE
#sns.kdeplot(df['O₃'], shade=True)
#sns.kdeplot(df['SO₂'], shade=True)

#box plot
#sns.boxplot(data=df, x="O₃", orient = 'v')
# sns.boxplot(data=df.iloc[:,6:8])

fig, ax = plt.subplots(1,2,figsize=(8,3))

sns.distplot(df['PM-10'], hist_kws={'range':[0,300]}, ax = ax[0])

sns.distplot(df['PM-10'], ax = ax[1])
ax[1].set_xscale('log')
ax[1].set_title('Log bins')

#plt.legend()
plt.show()

quit()

sheet_names = pd.ExcelFile("result_statistics.xlsx").sheet_names

dfs = {}

for sheet in sheet_names:
        dfs[sheet] = pd.read_excel("result_statistics.xlsx", sheet)

ws = dfs['average']

site_name = ws['Unnamed: 0']

sns.barplot(
        data = ws,
        x = site_name,
        y = "PM-10"
)

# sns.distplot(ws['PM-10'])

# sns.distplot(ws['PM-2.5'])

#sns.relplot(x = "O₃", y = "PM-10", data = ws)

plt.xticks(rotation = -45)
plt.show()


#ws = rb['average'] #min max

#print(ws.head())