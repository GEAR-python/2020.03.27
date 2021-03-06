import openpyxl as pyxl
import pandas as pd
import numpy as np

import datetime
from dateutil.parser import parse


#프로그램 내에서 사용할 list 정의 (생성)

time = list()
pol_data = list()
date = list()
tot_date = list()
O3 = list()
NO = list()
CO = list()
SO = list()
PM10 = list()
PM25 = list()

#프로그램 내에서 사용할 dictionary 정의 (생성)
dic_mean = {}
dic_max = {}
dic_min = {}


#관측 지점 목록
obs_site = ['광복동', '장림동', '학장동', '덕천동', '연산동', '대연동', '청룡동', 
     '전포동', '태종대','기장읍','대저동','부곡동','광안동','명장동','녹산동','용수리',
         '좌동','수정동','대신동','온천동','초량동']

# test 
#obs_site = ['광복동', '장림동'] # test

#일(day) 수 초기화
num_day = 0

#24시간 list를 미리 만들기
#10 이하 시간에는 숫자'0'추가, 10 이상에는 그냥 time 리스트에 저장
for t in range(0,24):
    if t < 10 :
        time.append('0'+ str(t))
    else:
        time.append(t)
# time.append((r[t].value)[:-1])


writer = pd.ExcelWriter('result_test.xlsx') #output파일 생성
# test # writer = pd.ExcelWriter('result_jeonpo.xlsx') #output파일 생성     


for site_name in obs_site: #지점별로 loop

    for yr in range(2010, 2011):# 2017):
        for mn in range(1, 13): #  13):

            if site_name == '학장동' and yr == 2010 and mn < 7 : #감전동 (구 학장동) 지점 교체된 것 고려
                site_name = '감전동'
            elif site_name == '감전동' and yr == 2010 and mn > 6:
                site_name = '학장동'

            file_name = "data/측정기록지("+str(yr)+"년"+str(mn)+"월).xlsx"

            print(site_name, yr, mn, file_name)

            wb = pyxl.load_workbook(file_name)

            if site_name == '수정동' or site_name == '대신동': #수정동, 대신동이 엑셀파일 sheet 중 없을 경우 break
                list_sheet = wb.get_sheet_names()
                if site_name not in list_sheet:
                    break

            #test = wb.worksheets[0] #시트 순서별로 불러오기
            test = wb.get_sheet_by_name(site_name)
            
            pol_data = list()
            # name = list()
            # time = list()
            date = list()

            #open한 엑셀 파일의 line 수 count
            check_line = 0

            for r in test.rows:
                check_line = check_line + 1

                if r[0].value == '측 정 소: ' :
                    continue
                elif r[0].value == '측정항목: ' :
                    name = r[2].value
                    print(name)
                elif r[0].value == '측정년월: ' :
                    year = (r[1].value)[0:5]
                    #month = int((r[2].value)[0:2])
                    month = (r[2].value)[0:2]
                   # print(year, month)
                elif r[0].value == '구 분' :
                    continue
                elif r[0].value == '최소' :
                    continue                
                elif r[0].value == '최대' :
                    if name == 'O₃':
                        O3 = O3 + pol_data
                        pol_data = list()
                        date = list()
                    elif name == 'SO₂':
                        SO = SO + pol_data
                        pol_data = list()
                        date = list()
                    elif name == 'CO':
                        CO = CO + pol_data
                        pol_data = list()
                        date = list()
                    elif name == 'NO₂':
                        NO = NO + pol_data
                        pol_data = list()
                        date = list()
                    elif name == 'PM-10':
                        if test.max_row == check_line + 1: # 마지막 행(row)일 경우 = PM2.5가 없을 경우
                            print(test.max_row, check_line)
                            PM10 = PM10 + pol_data
                            tot_date = tot_date + date

                            #PM25 리스트에 일수만큼의 null값(-999, Nan값)을 넣어라 (일수*24시간)
                            for x in range(num_day*24):
                                #PM25.append(-999)
                                PM25.append(np.nan)
                        else:  #마지막 행이 아닐경우 = 뒤에 PM2.5가 남았을 경우
                            PM10 = PM10 + pol_data
                            print(test.max_row, check_line)
                            pol_data = list()
                            date = list()
                    elif name == 'PM-2.5':
                        PM25 = PM25 + pol_data
                        tot_date = tot_date + date
                        break
                    elif name == '온도':
                        tot_date = tot_date + date

                        for x in range(num_day*24):
                                #PM25.append(-999)
                                PM25.append(np.nan)

                        break
                    else:
                        pol_data = list()
                        date = list()
                elif r[0].value == '평균' :
                    num_day = 0
                    continue
                elif r[0].value == None :
                    continue
                else : #오염 자료가 있는 1일부터 말일까지 읽기
                    #날짜 중 '일'에 숫자 0 붙히기
                    num_day = num_day + 1

                    if int((r[0].value)[:-1]) < 10 :
                        day = '0'+ (r[0].value)[:-1]
                    else :
                        day = (r[0].value)[:-1]

                    #각 오염 농도값을 pol_data 리스트에 추가하기
                    for i in range(1,25):
                        obj = r[i].value

                        #자료 중 '교정중, 동작불량 등 문자열들을 nan 값으로 치환
                        if str(type(r[i].value)) == "<class 'str'>":
                            #print(type(r[i].value))
                            obj = np.nan

                        pol_data.append(obj)

                        ## 연월일시간에다가 분은 00으로 설정. parse는 시간만 있고 분이 없으니깐 인식을 못함
                        ## parse 이용해서 날짜+시간의 형식 변형
                        conv_date = str(year)+str(month)+str(day)+str(time[i-1])+'00'

                        date.append(parse(conv_date))

                        #date.append(datetime.datetime.strptime(conv_date,"%Y%m%d%H"))

                      #  date.append(str(year)+str(month)+str(day)+str(time[i-1]))
                     
    #지점별, 오염별 자료를 DataFrame 형태로 작성
    df = pd.DataFrame({'Date':pd.Series(tot_date),'O₃':pd.Series(O3), 'SO₂':pd.Series(SO), 'CO':pd.Series(CO), 'NO₂':pd.Series(NO),'PM-10':pd.Series(PM10), 'PM-2.5':pd.Series(PM25)})
    
    print(df)

    df.to_excel(writer, sheet_name=site_name) #엑셀 결과로 저장
   
   #각 오염별 평균, 최대, 최소값을 계산하고 정의해둔 dictionary에 저장
    mean_value = df.mean(skipna=True)

    max_value = df.max(skipna=True)

    min_value = df.min(skipna=True)

    dic_mean[site_name] = mean_value

    dic_max[site_name] = max_value

    dic_min[site_name] = min_value

    writer.save()

    #list들 초기화
    tot_date = list()
    O3 = list()
    NO = list()
    CO = list()
    SO = list()
    PM10 = list()
    PM25 = list()

# index를 date로 바꾸려면 아래 문장 사용해야함.
#df.set_index('date', inplace=True)
#print(df)


#통계치 계산해서 저장할 결과 파일 만들기

list_pol = ['O₃', 'SO₂', 'CO', 'NO₂','PM-10','PM-2.5']

result_read = pyxl.Workbook()

ws_mean = result_read.create_sheet('average',1)
ws_max =  result_read.create_sheet('max',2)
ws_min =  result_read.create_sheet('min',3)

n = 1

for data in list_pol:
    ws_mean.cell(row=1, column=n+1).value = list_pol[n-1]
    ws_max.cell(row=1, column=n+1).value = list_pol[n-1]
    ws_min.cell(row=1, column=n+1).value = list_pol[n-1]

    n = n + 1

m = 2
o = 1

for name in obs_site:
    ws_mean.cell(row=m, column=1).value = name
    ws_max.cell(row=m, column=1).value = name
    ws_min.cell(row=m, column=1).value = name
    for data in list_pol:
        ws_mean.cell(row=m, column=o+1).value = dic_mean[name][data]
        ws_max.cell(row=m, column=o+1).value = dic_max[name][data]
        ws_min.cell(row=m, column=o+1).value = dic_min[name][data]

        o = o + 1
    o = 1

    m = m + 1
    
result_read.save('result_statistics.xlsx')
